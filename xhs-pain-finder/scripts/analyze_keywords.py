#!/usr/bin/env python3
"""
词频与情感分析脚本
分析评论高频词、情感倾向，生成分析报告
"""

import argparse
import json
import re
from pathlib import Path
from collections import Counter
from datetime import datetime

def analyze_keywords(json_path: str, output_path: str, top_n: int = 50):
    """
    分析评论词频和情感
    
    Args:
        json_path: 输入 JSON 文件路径
        output_path: 输出 Excel 文件路径
        top_n: TOP N 高频词
    """
    # 检查依赖
    try:
        import jieba
        import jieba.analyse
    except ImportError:
        print("错误: 请先安装 jieba: pip install jieba")
        return None
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("错误: 请先安装 openpyxl: pip install openpyxl")
        return None
    
    try:
        from snownlp import SnowNLP
        has_snownlp = True
    except ImportError:
        print("警告: 未安装 snownlp，跳过情感分析")
        has_snownlp = False
    
    # 读取数据
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    comments = data.get("comments", [])
    if not comments:
        print("警告: 没有找到评论数据")
        return None
    
    print(f"正在分析 {len(comments)} 条评论...")
    
    # 合并所有评论文本
    all_text = "\n".join([c.get("content", "") for c in comments])
    
    # 停用词
    stopwords = set([
        "的", "了", "是", "我", "你", "他", "她", "它", "们", "这", "那",
        "有", "在", "不", "也", "就", "都", "要", "会", "很", "到", "说",
        "还", "能", "对", "和", "与", "吗", "吧", "啊", "呢", "哦", "嗯",
        "什么", "怎么", "为什么", "哪", "哪里", "这个", "那个", "一个",
        "可以", "没有", "因为", "所以", "但是", "如果", "虽然", "而且",
        "或者", "以及", "比如", "就是", "不是", "可能", "应该", "觉得",
        "知道", "看到", "感觉", "真的", "确实", "其实", "然后", "已经"
    ])
    
    # 分词统计
    words = jieba.cut(all_text)
    word_counts = Counter()
    
    for word in words:
        word = word.strip()
        if len(word) >= 2 and word not in stopwords and not word.isdigit():
            if not re.match(r'^[\W_]+$', word):  # 排除纯标点
                word_counts[word] += 1
    
    top_words = word_counts.most_common(top_n)
    
    # 提取关键词（TF-IDF）
    keywords_tfidf = jieba.analyse.extract_tags(all_text, topK=20, withWeight=True)
    
    # 情感分析
    sentiments = []
    if has_snownlp:
        for comment in comments:
            content = comment.get("content", "")
            if content:
                try:
                    s = SnowNLP(content)
                    score = s.sentiments  # 0-1, 越大越正面
                    if score > 0.6:
                        sentiment = "正面"
                    elif score < 0.4:
                        sentiment = "负面"
                    else:
                        sentiment = "中性"
                    sentiments.append({
                        "content": content[:100],
                        "score": round(score, 3),
                        "sentiment": sentiment,
                        "likes": comment.get("likes", 0)
                    })
                except:
                    pass
    
    # 统计情感分布
    sentiment_counts = Counter([s["sentiment"] for s in sentiments])
    
    # 提取痛点关键词（负面评论中的高频词）
    pain_words = Counter()
    if sentiments:
        negative_texts = [s["content"] for s in sentiments if s["sentiment"] == "负面"]
        for text in negative_texts:
            for word in jieba.cut(text):
                word = word.strip()
                if len(word) >= 2 and word not in stopwords:
                    pain_words[word] += 1
    
    # 创建 Excel
    wb = Workbook()
    
    # ===== Sheet 1: 高频词统计 =====
    ws1 = wb.active
    ws1.title = "高频词TOP50"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    
    headers1 = ["排名", "词语", "出现次数", "词频占比"]
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    total_words = sum(word_counts.values())
    for row, (word, count) in enumerate(top_words, 2):
        ws1.cell(row=row, column=1, value=row - 1)
        ws1.cell(row=row, column=2, value=word)
        ws1.cell(row=row, column=3, value=count)
        ws1.cell(row=row, column=4, value=f"{count/total_words*100:.2f}%")
    
    ws1.column_dimensions["B"].width = 20
    
    # ===== Sheet 2: TF-IDF 关键词 =====
    ws2 = wb.create_sheet("TF-IDF关键词")
    
    headers2 = ["排名", "关键词", "权重"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    for row, (word, weight) in enumerate(keywords_tfidf, 2):
        ws2.cell(row=row, column=1, value=row - 1)
        ws2.cell(row=row, column=2, value=word)
        ws2.cell(row=row, column=3, value=round(weight, 4))
    
    ws2.column_dimensions["B"].width = 20
    
    # ===== Sheet 3: 情感分析 =====
    if sentiments:
        ws3 = wb.create_sheet("情感分析")
        
        # 汇总
        ws3.cell(row=1, column=1, value="情感分布汇总").font = Font(bold=True)
        ws3.cell(row=2, column=1, value="正面")
        ws3.cell(row=2, column=2, value=sentiment_counts.get("正面", 0))
        ws3.cell(row=3, column=1, value="中性")
        ws3.cell(row=3, column=2, value=sentiment_counts.get("中性", 0))
        ws3.cell(row=4, column=1, value="负面")
        ws3.cell(row=4, column=2, value=sentiment_counts.get("负面", 0))
        
        # 明细
        headers3 = ["评论内容", "情感得分", "情感类型", "点赞数"]
        for col, header in enumerate(headers3, 1):
            cell = ws3.cell(row=6, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # 按点赞数排序
        sorted_sentiments = sorted(sentiments, key=lambda x: x["likes"], reverse=True)
        for row, s in enumerate(sorted_sentiments[:100], 7):
            ws3.cell(row=row, column=1, value=s["content"])
            ws3.cell(row=row, column=2, value=s["score"])
            ws3.cell(row=row, column=3, value=s["sentiment"])
            ws3.cell(row=row, column=4, value=s["likes"])
            
            # 根据情感类型着色
            if s["sentiment"] == "负面":
                ws3.cell(row=row, column=3).fill = PatternFill("solid", fgColor="FFC7CE")
            elif s["sentiment"] == "正面":
                ws3.cell(row=row, column=3).fill = PatternFill("solid", fgColor="C6EFCE")
        
        ws3.column_dimensions["A"].width = 60
    
    # ===== Sheet 4: 痛点关键词 =====
    if pain_words:
        ws4 = wb.create_sheet("痛点关键词")
        
        headers4 = ["排名", "痛点词", "出现次数"]
        for col, header in enumerate(headers4, 1):
            cell = ws4.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row, (word, count) in enumerate(pain_words.most_common(30), 2):
            ws4.cell(row=row, column=1, value=row - 1)
            ws4.cell(row=row, column=2, value=word)
            ws4.cell(row=row, column=3, value=count)
        
        ws4.column_dimensions["B"].width = 20
    
    # 保存
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    
    print(f"\n✅ 分析完成！")
    print(f"📄 保存至: {output_file}")
    print(f"\n📊 分析结果摘要:")
    print(f"   - 高频词 TOP5: {', '.join([w for w, c in top_words[:5]])}")
    print(f"   - TF-IDF 关键词: {', '.join([w for w, _ in keywords_tfidf[:5]])}")
    if sentiments:
        print(f"   - 情感分布: 正面 {sentiment_counts.get('正面', 0)}, 中性 {sentiment_counts.get('中性', 0)}, 负面 {sentiment_counts.get('负面', 0)}")
    if pain_words:
        print(f"   - 痛点词 TOP5: {', '.join([w for w, c in pain_words.most_common(5)])}")
    
    return {
        "top_words": top_words,
        "keywords_tfidf": keywords_tfidf,
        "sentiment_counts": dict(sentiment_counts),
        "pain_words": pain_words.most_common(30)
    }


def main():
    parser = argparse.ArgumentParser(description="评论词频与情感分析")
    parser.add_argument("json_file", help="输入 JSON 文件路径")
    parser.add_argument("--output", "-o", default="分析结果.xlsx", help="输出 Excel 文件路径")
    parser.add_argument("--top", type=int, default=50, help="TOP N 高频词")
    
    args = parser.parse_args()
    analyze_keywords(args.json_file, args.output, args.top)


if __name__ == "__main__":
    main()
