#!/usr/bin/env python3
"""
评论数据存储脚本 - 将 JSON 数据保存到 Excel
"""

import argparse
import json
from pathlib import Path
from datetime import datetime

def save_to_excel(json_path: str, output_path: str):
    """
    将评论 JSON 数据保存到 Excel
    
    Args:
        json_path: 输入 JSON 文件路径
        output_path: 输出 Excel 文件路径
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("错误: 请先安装 openpyxl: pip install openpyxl")
        return None
    
    # 读取 JSON 数据
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    comments = data.get("comments", [])
    
    if not comments:
        print("警告: 没有找到评论数据")
        return None
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "评论数据"
    
    # 样式定义
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # 表头
    headers = ["序号", "用户昵称", "评论内容", "点赞数", "发布时间", "作者回复", "子评论"]
    col_widths = [8, 15, 60, 10, 15, 10, 40]
    
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    # 填充数据
    for row_idx, comment in enumerate(comments, 2):
        row_data = [
            comment.get("index", row_idx - 1),
            comment.get("nickname", ""),
            comment.get("content", ""),
            comment.get("likes", 0),
            comment.get("time", ""),
            "是" if comment.get("is_author_reply") else "",
            "\n".join(comment.get("sub_comments", []))
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = cell_align
            cell.border = border
        
        # 根据点赞数高亮
        likes = comment.get("likes", 0)
        if likes >= 100:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="FFF2CC")
        elif likes >= 50:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="E2EFDA")
    
    # 添加汇总信息工作表
    ws_summary = wb.create_sheet("汇总信息")
    
    summary_data = [
        ["帖子链接", data.get("url", "")],
        ["帖子标题", data.get("title", "")],
        ["抓取时间", data.get("crawl_time", "")],
        ["评论总数", len(comments)],
        ["总点赞数", sum(c.get("likes", 0) for c in comments)],
        ["作者回复数", sum(1 for c in comments if c.get("is_author_reply"))],
        ["有子评论数", sum(1 for c in comments if c.get("sub_comments"))],
    ]
    
    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=2, value=value)
    
    ws_summary.column_dimensions["A"].width = 15
    ws_summary.column_dimensions["B"].width = 80
    
    # 保存文件
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    
    print(f"✅ Excel 文件已保存: {output_file}")
    print(f"   - 共 {len(comments)} 条评论")
    print(f"   - 高赞评论(≥100): {sum(1 for c in comments if c.get('likes', 0) >= 100)} 条")
    
    return output_file


def main():
    parser = argparse.ArgumentParser(description="将评论数据保存到 Excel")
    parser.add_argument("json_file", help="输入 JSON 文件路径")
    parser.add_argument("--output", "-o", default="评论数据.xlsx", help="输出 Excel 文件路径")
    
    args = parser.parse_args()
    save_to_excel(args.json_file, args.output)


if __name__ == "__main__":
    main()
